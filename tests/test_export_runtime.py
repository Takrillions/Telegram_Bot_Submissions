import csv
from io import BytesIO, TextIOWrapper
import unittest
from openpyxl import load_workbook
from export_runtime import SHEETS, csv_export, xlsx_export


def snapshot():
    return {"channel_id": 7, "period": "30d", "metadata": {"channel_title": "=unsafe, name", "timezone": "UTC", "generated_at": "2026-08-14T00:00:00+00:00"}, "conversation_metrics_complete": False, "statistics": {"unique_subscribers": 0, "media": {"text": 1, "photo": 0}, "messages_by_hour": {0: 1}, "messages_by_weekday": {0: 1}, "top_subscribers": [{"display_name": "=formula", "message_count": 2}]}, "administrators": {"tracked_conversation_count": 2, "handled_conversation_count": 1, "unanswered_conversation_count": 1, "admins": [{"admin_id": 1, "display_name": "@admin", "reply_count": 1, "unique_conversations_replied": 1, "handled_conversations": 1, "first_response_count": 1, "average_first_response_seconds": 10, "median_first_response_seconds": 10, "moderation_actions": 0, "restrictions": 0, "warnings": 0, "spam_marks": 0, "management_actions": 0}]}}

class ExportRuntimeTests(unittest.TestCase):
    def test_csv_bom_quoting_and_formula_safety(self):
        data=csv_export(snapshot()); self.assertTrue(data.startswith(b'\xef\xbb\xbf'))
        rows=list(csv.reader(TextIOWrapper(BytesIO(data),encoding='utf-8-sig')))
        self.assertEqual(rows[0],["section","metric","value"])
        self.assertIn(["top_subscriber","'=formula","2"],rows)
        self.assertIn(["metadata","channel_title","'=unsafe, name"],rows)
        self.assertIn(["administration","unanswered_conversation_count","1"],rows)
        self.assertIn(["administrator","'@admin:handled_conversations","1"],rows)
    def test_xlsx_sheets_and_formula_safety(self):
        book=load_workbook(BytesIO(xlsx_export(snapshot())))
        self.assertEqual(tuple(book.sheetnames),SHEETS)
        self.assertEqual(book[SHEETS[4]]["A2"].value,"'=formula")
        self.assertEqual(book[SHEETS[0]]["B2"].value,"'=unsafe, name")
        self.assertEqual(book[SHEETS[5]]["B4"].value,1)
        self.assertIn("handled_conversations", [cell.value for cell in book[SHEETS[5]][6]])

if __name__ == '__main__': unittest.main()
